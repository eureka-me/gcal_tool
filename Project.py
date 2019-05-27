#! env python
# -*- coding: utf-8 -*-
# Date: 2018/11/05
# Filename: Project 

__author__ = 'kke'
__date__ = "2018/11/05"

import configparser
import xlrd
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

from module import GetGcalInfo as ggi
from module import ProcessGcalData as pgd


config = configparser.ConfigParser()
config.read_file(open('./config.conf', 'r', encoding='UTF-8'))

GetGcalInfo = ggi.GetGcalInfo(config)


def main():

    # トレース対象プロジェクトの取得（エクセルを読み込んで作成）
    projects, abbrs, time_min = defaultdict(lambda: defaultdict(dict)), [], datetime(2100, 1, 1)
    proj_file_dir = config['GENERAL']['PROJECT_DIR']
    proj_book = xlrd.open_workbook(proj_file_dir + 'src.xlsx')
    for sheet_name in proj_book.sheet_names():
        if sheet_name == 'prj名':
            continue
        else:
            proj_sheet = proj_book.sheet_by_name(sheet_name)
            row_dic = {proj_sheet.cell(i, 0).value: i for i in range(proj_sheet.nrows)}
            st_date = excel_date(proj_sheet.cell(row_dic['開始日'], 1).value)
            en_date = excel_date(proj_sheet.cell(row_dic['終了日'], 1).value)
            projects[proj_sheet.cell(row_dic['prj略記'], 1).value]\
                [(st_date, en_date)]['prj名'] = \
                proj_sheet.cell(row_dic['prj名'], 1).value

            projects[proj_sheet.cell(row_dic['prj略記'], 1).value] \
                [(st_date, en_date)]['実績'] = \
                row_dic['実績']

            projects[proj_sheet.cell(row_dic['prj略記'], 1).value] \
                [(st_date, en_date)]['sheet_name'] = \
                sheet_name

            abbrs.append(proj_sheet.cell(row_dic['prj略記'], 1).value)

            if st_date < time_min:
                time_min = st_date

    ### 集計
    # もっとも古い日時から取ってくる
    time_max = datetime.now()
    work_result_info, _ = GetGcalInfo.get_gcal_info(config['RETRIEVE']['WORK_RESULT_CAL_ID'], time_min=time_min, time_max=time_max)
    PGD = pgd.ProcessGcalData(config, None)
    result_dic = PGD.summarize_calendar_info(calendar_info=work_result_info, cate_scope=10, dic=defaultdict(float))

    # 週ごとに集計
    for (abbr, date_str), _min in result_dic.items():
        if abbr in abbrs:
            _date = datetime.strptime(date_str, '%Y-%m-%d')
            for (st_date, en_date) in projects[abbr].keys():
                if st_date <= _date <= en_date:
                    if 'result' not in projects[abbr][(st_date, en_date)]:
                        print(1)
                        projects[abbr][(st_date, en_date)]['result'] = defaultdict(int)
                    projects[abbr][(st_date, en_date)]['result'][_date] += _min
                    print(2)

    # excelファイルの上書き
    wb = openpyxl.load_workbook(proj_file_dir + 'src.xlsx')
    for abbr, _dic in projects.items():
        for (st_date, en_date), _dic2 in _dic.items():
            result_nrow = _dic2['実績']

            # 列の位置を返す辞書
            i, col_dic = 0, {}
            st_week = st_date - timedelta(days=st_date.weekday())
            for i in range(30):
                col_dic[st_week + timedelta(days=i * 7)] = {'col': 4 + i, 'time': 0}
                i += 1

            for d, t in _dic2['result'].items():
                d_w = d - timedelta(days=d.weekday())
                col_dic[d_w]['time'] += t / 60

            for d_w, _dic3 in col_dic.items():
                wb[_dic2['sheet_name']].cell(row=result_nrow + 1, column=_dic3['col'] + 1, value=_dic3['time'])

    wb.save(proj_file_dir + 'src.xlsx')


def excel_date(num):
    return datetime(1899, 12, 30) + timedelta(days=num)


if __name__ == '__main__':
    main()
